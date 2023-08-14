import uuid
from itertools import count
from uuid import UUID

from openpyxl import Workbook, load_workbook
from src.api.v1.dishes.schemas import Dish
from src.api.v1.menus.schemas import Menu
from src.api.v1.submenus.schemas import Submenu


def parse_dish(row, submenu_id: UUID):
    id = row[2].value
    if id is None:
        return None

    name = row[3].value
    desc = row[4].value
    price = str(row[5].value)

    # create uuid from menu id in table, so it should not be changed
    id = uuid.uuid5(namespace=uuid.NAMESPACE_OID, name=f'{submenu_id}_{id}')
    return Dish.model_validate({'id': id, 'title': name, 'description': desc, 'submenu_id': submenu_id, 'price': price})


def parse_submenu(row, menu_id: UUID):
    id = row[1].value
    if id is None:
        return None

    name = row[2].value
    desc = row[3].value

    # create uuid from menu id in table, so it should not be changed
    id = uuid.uuid5(namespace=uuid.NAMESPACE_OID, name=f'{menu_id}_{id}')
    return Submenu.model_validate({'id': id, 'title': name, 'description': desc, 'menu_id': menu_id})


def parse_menu(row):
    id = row[0].value
    if id is None:
        return None

    name = row[1].value
    desc = row[2].value

    # create uuid from menu id in table, so it should not be changed
    id = uuid.uuid5(namespace=uuid.NAMESPACE_OID, name=str(id))
    return Menu.model_validate({'id': id, 'title': name, 'description': desc})


def parse_wb() -> tuple[list[Menu], list[Submenu], list[Dish]]:
    wb: Workbook = load_workbook('admin/Menu.xlsx')
    ws = wb.active
    menus = []
    submenus = []
    dishes = []
    for row_index in count(start=1, step=1):
        row = ws[f'A{row_index}:F{row_index}'][0]
        menu = parse_menu(row)
        if menu is not None:
            menus.append(menu)
            continue

        # if no menus found, ignore wrong written line
        if len(menus) == 0:
            continue

        submenu = parse_submenu(row, menus[-1].id)
        if submenu is not None:
            submenus.append(Submenu.model_validate(submenu))
            continue

        # if no submenus found, ignore wrong written line
        if len(submenus) == 0:
            continue

        dish = parse_dish(row, submenus[-1].id)
        if dish is not None:
            dishes.append(Dish.model_validate(dish))
        else:
            break

    return menus, submenus, dishes


parse_wb()
